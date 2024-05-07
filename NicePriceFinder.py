from HelperFuncitons import *
from PositionVariables import *
from watchdog import setup_watchdog
import openpyxl
market1_prices = []


class OrderInfo:
    def __init__(self, item_name, sell_price, buy_price):
        self.item_name = item_name
        self.sell_price = sell_price
        self.buy_price = buy_price
        self.price_difference = ((sell_price - buy_price) / buy_price) * 100 if buy_price else 0

    def __repr__(self):
        return f"{self.item_name}: Sellorder: {self.sell_price}, Buyorder: {self.buy_price}"


def scan_collect_price(item_name, market_prices): #START THIS IN BUY SECTION
    print("checking item: "+item_name)
    positions = PositionConfig()
    click(positions.ITEM_NAME_ENTRY_POS)
    type_with_delay(item_name, 0.0001)
    time.sleep(0.3)
    click(positions.BUY_BUTTON_POS)
    check_open_orderoverview(positions)
    time.sleep(1)
    sell_price, buy_price = scan_top_orders(positions)
    market_prices.append(OrderInfo(item_name, sell_price, buy_price))
    click(positions.CLOSE_BUTTON)


def sort_prices():
    # Sortieren der Liste nach prozentualem Unterschied, absteigend
    market1_prices.sort(key=lambda x: x.price_difference, reverse=True)

    # Ausgabe der sortierten Liste
    for order in market1_prices:
        print(order)


items = ["Attentäterjacke", "Soldatenrüstung"]

def findNicePrice():
    for item in items:
        scan_collect_price(item, market1_prices)

    sort_prices()


def update_excel_with_prices(excel_file_path, market_prices):
    # Excel-Datei öffnen
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active  # Gehe davon aus, dass wir im ersten Tabellenblatt arbeiten

    # Durch jede Preisinformation in market_prices iterieren
    for order_info in market_prices:
        for row in range(2, sheet.max_row + 1):  # Starte bei 2, falls Zeile 1 Kopfzeilen enthält
            # Durchsuche die relevanten Spalten nach dem Gegenstandsnamen
            for col in [1, 7, 13]:  # Spalten 1, 7, 13 nach dem Namen durchsuchen
                item_cell = sheet.cell(row=row, column=col)
                if item_cell.value == order_info.item_name:
                    # Preis 2 Felder rechts von der gefundenen Zelle schreiben
                    sheet.cell(row=row, column=col + 2, value=order_info.sell_price)  # Verkaufspreis 2 Felder rechts
                    break  # Beende die Schleife, wenn der Gegenstandsname gefunden wurde

    # Speichere die Änderungen in der Excel-Datei
    workbook.save(excel_file_path)

# Am Ende deines Hauptprogramms rufe diese Funktion auf
if __name__ == "__main__":
    setup_watchdog()
    findNicePrice()
    update_excel_with_prices('C:\\Users\\Louis\\Desktop\\Albion_BM_Profit_sheet_2.xlsx', market1_prices)